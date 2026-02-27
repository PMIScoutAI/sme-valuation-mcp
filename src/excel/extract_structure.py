from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils.cell import (
    column_index_from_string,
    coordinate_from_string,
    get_column_letter,
)

# Optional sheet name + A1 ref, with optional range A1:B2.
CELL_REF_RE = re.compile(
    r"(?:(?P<sheet>'[^']+'|[A-Za-z0-9_.]+)!)?"
    r"(?P<start>\$?[A-Z]{1,3}\$?\d+)"
    r"(?::(?P<end>\$?[A-Z]{1,3}\$?\d+))?"
)


@dataclass(frozen=True)
class FormulaCell:
    address: str
    formula: str


def normalize_sheet_name(raw: str | None) -> str | None:
    if raw is None:
        return None
    cleaned = raw.strip()
    if cleaned.startswith("'") and cleaned.endswith("'"):
        cleaned = cleaned[1:-1].replace("''", "'")
    return cleaned


def strip_dollar(coord: str) -> str:
    return coord.replace("$", "")


def expand_range(start: str, end: str, max_cells: int = 10000) -> Iterable[str]:
    start_col, start_row = coordinate_from_string(strip_dollar(start))
    end_col, end_row = coordinate_from_string(strip_dollar(end))

    min_col = min(column_index_from_string(start_col), column_index_from_string(end_col))
    max_col = max(column_index_from_string(start_col), column_index_from_string(end_col))
    min_row = min(start_row, end_row)
    max_row = max(start_row, end_row)

    total = (max_col - min_col + 1) * (max_row - min_row + 1)
    if total > max_cells:
        # Safety guard: avoid exploding very large references.
        return [strip_dollar(start), strip_dollar(end)]

    refs: list[str] = []
    for c_idx in range(min_col, max_col + 1):
        col = get_column_letter(c_idx)
        for r_idx in range(min_row, max_row + 1):
            refs.append(f"{col}{r_idx}")
    return refs


def extract_local_refs(formula: str, current_sheet: str) -> set[str]:
    refs: set[str] = set()
    for match in CELL_REF_RE.finditer(formula):
        sheet = normalize_sheet_name(match.group("sheet"))
        if sheet is not None and sheet != current_sheet:
            continue

        start = match.group("start")
        end = match.group("end")
        if not start:
            continue

        if end:
            refs.update(expand_range(start, end))
        else:
            refs.add(strip_dollar(start))
    return refs


def is_formula(value: object, data_type: str) -> bool:
    return data_type == "f" or (isinstance(value, str) and value.startswith("="))


def sheet_formula_cells(ws) -> list[FormulaCell]:
    formulas: list[FormulaCell] = []
    for row in ws.iter_rows():
        for cell in row:
            if is_formula(cell.value, cell.data_type):
                formula_value = str(cell.value)
                if not formula_value.startswith("="):
                    formula_value = f"={formula_value}"
                formulas.append(FormulaCell(address=cell.coordinate, formula=formula_value))
    return formulas


def sort_coord(coord: str) -> tuple[int, int]:
    col, row = coordinate_from_string(coord)
    return row, column_index_from_string(col)


def build_report(excel_path: Path, top_n: int) -> dict:
    wb = load_workbook(excel_path, data_only=False, read_only=False)

    sheets_report: list[dict] = []
    candidates_report: dict[str, list[dict]] = {}

    for ws in wb.worksheets:
        formulas = sheet_formula_cells(ws)
        referenced_by_formulas: set[str] = set()
        formula_by_cell = {f.address: f.formula for f in formulas}

        for formula_cell in formulas:
            referenced_by_formulas.update(extract_local_refs(formula_cell.formula, ws.title))

        terminal_cells = [
            fc.address
            for fc in formulas
            if strip_dollar(fc.address) not in referenced_by_formulas
        ]
        terminal_cells = sorted(terminal_cells, key=sort_coord)

        candidates_report[ws.title] = [
            {
                "cell": address,
                "formula": formula_by_cell[address],
            }
            for address in terminal_cells[:top_n]
        ]

        sheets_report.append(
            {
                "sheet": ws.title,
                "used_range": ws.calculate_dimension(),
                "formula_count": len(formulas),
                "terminal_candidates_count": len(terminal_cells),
            }
        )

    return {
        "source_file": str(excel_path),
        "sheets": sheets_report,
        "candidates": candidates_report,
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Extract workbook structure and terminal formula candidates."
    )
    parser.add_argument("--excel", required=True, help="Path to source Excel workbook")
    parser.add_argument(
        "--output",
        default="spec/structure_report.json",
        help="Output JSON path",
    )
    parser.add_argument(
        "--top-n",
        type=int,
        default=30,
        help="Top N terminal candidates per sheet",
    )
    args = parser.parse_args()

    excel_path = Path(args.excel).resolve()
    output_path = Path(args.output).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    report = build_report(excel_path=excel_path, top_n=args.top_n)
    output_path.write_text(json.dumps(report, indent=2), encoding="utf-8")

    print(f"Report written to: {output_path}")


if __name__ == "__main__":
    main()
